#!/usr/bin/env python3
"""
Extract the SAMO Team seed from externaldata/roledata.xlsx + previousroledata.json.

- Role TREE comes from previousroledata.json (division -> department -> roles),
  preserving its order (the user's authoritative hierarchy/ordering source).
- Any xlsx role (ตำแหน่ง 1-3) that doesn't match a json role is attached as a
  loose role node under its division (nothing is lost). RT/MDI divisions have no
  json roles, so all their roles land this way — expected.
- All 340 people are imported onto their matched role node (one member row per
  person-role), carrying ยืนยัน (confirmed).

Outputs:
  externaldata/team-seed.json           (canonical, inspectable)
  supabase/migrations/0047_seed_team_data.sql  (generated inserts)

Run:  python3 tools/extract-team-seed.py
"""
import json, re, uuid, os, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'externaldata', 'roledata.xlsx')
PREV = os.path.join(ROOT, 'externaldata', 'previousroledata.json')
OUT_JSON = os.path.join(ROOT, 'externaldata', 'team-seed.json')
OUT_SQL = os.path.join(ROOT, 'supabase', 'migrations', '0047_seed_team_data.sql')

import openpyxl

# sheet tab name -> json division name
SHEET_DIVISION = {
    'บริหาร': 'ฝ่ายบริหารองค์กร',
    'ดิจิมอน': 'ฝ่ายดิจิทัลและสื่อสารองค์กร',
    'ใน': 'ฝ่ายกิจการภายใน',
    'นอก': 'ฝ่ายกิจการภายนอก',
    'มหาลัย': 'ฝ่ายกิจการมหาวิทยาลัย',
    'วิชาการ': 'ฝ่ายวิชาการ',
    'ยุทธ์ฯ': 'ฝ่ายยุทธศาสตร์และพัฒนาองค์กร',
    'คุณภาพฯ': 'ฝ่ายคุณภาพชีวิตและสิ่งแวดล้อม',
    'RT': 'ฝ่ายรังสีเทคนิค',
    'MDI': 'ฝ่ายเวชนิทัศน์',
}


def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').strip())


def newid():
    return str(uuid.uuid4())


# ---- 1. Build the tree from previousroledata.json --------------------------
with open(PREV, encoding='utf-8') as f:
    prev = json.load(f)

nodes = []          # {id, parent_id, name, kind, position}
division_node = {}  # division name -> node id
# division name -> { normalized role name -> node id }  for matching
division_role_index = {}


def add_node(parent_id, name, kind, position):
    nid = newid()
    nodes.append({'id': nid, 'parent_id': parent_id, 'name': norm(name),
                  'kind': kind, 'position': position})
    return nid


for di, div in enumerate(prev):
    dname = norm(div['division'])
    div_id = add_node(None, dname, 'division', di)
    division_node[dname] = div_id
    division_role_index.setdefault(dname, {})
    for depi, dep in enumerate(div.get('departments', []) or []):
        dep_id = add_node(div_id, dep['name'], 'department', depi)
        roles = dep.get('roles', []) or []
        for ri, role in enumerate(roles):
            r_id = add_node(dep_id, role, 'role', ri)
            division_role_index[dname][norm(role)] = r_id

# Track next sibling position for loose nodes appended under a division.
division_next_pos = {d: len([n for n in nodes if n['parent_id'] == nid])
                     for d, nid in division_node.items()}


# ---- 2. Walk the xlsx, attach people + loose roles -------------------------
wb = openpyxl.load_workbook(XLSX, data_only=True)
members = []         # {id, node_id, position, ...}
node_member_pos = {} # node id -> next member position
unmatched_roles = {} # (division, role) -> node id  (created loose)
stats = {'people': 0, 'member_rows': 0, 'loose_nodes': 0}


def member_position(node_id):
    p = node_member_pos.get(node_id, 0)
    node_member_pos[node_id] = p + 1
    return p


def loose_role_node(dname, role):
    key = (dname, norm(role))
    if key in unmatched_roles:
        return unmatched_roles[key]
    div_id = division_node[dname]
    pos = division_next_pos.get(dname, 0)
    division_next_pos[dname] = pos + 1
    nid = add_node(div_id, role, 'role', pos)
    division_role_index[dname][norm(role)] = nid
    unmatched_roles[key] = nid
    stats['loose_nodes'] += 1
    return nid


def resolve_role(dname, role):
    rn = norm(role)
    idx = division_role_index.get(dname, {})
    if rn in idx:
        return idx[rn]
    return loose_role_node(dname, role)


for sheet, dname in SHEET_DIVISION.items():
    ws = wb[sheet]
    for r in range(3, ws.max_row + 1):
        mail = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value  # prefix OR full "prefix+name"
        c3 = ws.cell(r, 3).value  # full name
        roles = []
        for ci, yi in ((42, 45), (43, 46), (44, 47)):
            rv = ws.cell(r, ci).value
            if rv and norm(rv) and norm(rv) != '-':
                conf = ws.cell(r, yi).value
                roles.append((norm(rv), bool(conf)))
        if not mail and not roles and not c3:
            continue
        if not roles:
            continue
        # name handling: MDI sheet packs prefix+name into col2 (col3 empty)
        prefix, full_name = None, None
        if c3 and norm(c3):
            prefix = norm(c2) if c2 else None
            full_name = norm(c3)
        else:
            raw = norm(c2)
            m = re.match(r'^(นางสาว|นาย|นาง)\s*(.*)$', raw)
            if m:
                prefix, full_name = m.group(1), m.group(2).strip()
            else:
                prefix, full_name = None, raw
        nickname = norm(ws.cell(r, 4).value) or None
        student_id = norm(ws.cell(r, 5).value) or None
        # ชั้นปี → bare number for consistency ("ปี 5" / "5" / 5.0 → "5")
        yr = ws.cell(r, 6).value
        if isinstance(yr, float):
            yr = str(int(yr))
        ym = re.search(r'\d+', str(yr or ''))
        year = ym.group(0) if ym else None
        major = norm(ws.cell(r, 7).value) or None
        stats['people'] += 1
        for role, confirmed in roles:
            node_id = resolve_role(dname, role)
            members.append({
                'id': newid(), 'node_id': node_id,
                'position': member_position(node_id),
                'kkumail': norm(mail) or None, 'prefix': prefix,
                'full_name': full_name or '(ไม่ระบุชื่อ)', 'nickname': nickname,
                'student_id': student_id, 'year': year, 'major': major,
                'confirmed': confirmed,
            })
            stats['member_rows'] += 1

# ---- 3. Emit JSON + SQL ----------------------------------------------------
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump({'nodes': nodes, 'members': members, 'stats': stats},
              f, ensure_ascii=False, indent=2)


def sql_str(v):
    if v is None:
        return 'null'
    if isinstance(v, bool):
        return 'true' if v else 'false'
    return "'" + str(v).replace("'", "''") + "'"


lines = [
    '-- 0047_seed_team_data.sql',
    '-- GENERATED by tools/extract-team-seed.py — do not hand-edit.',
    '-- Seeds public.team_nodes (role tree) + public.team_members (people).',
    '-- Idempotent: truncates both tables first so a re-run replaces the seed.',
    '-- Tree order follows externaldata/previousroledata.json; people + loose',
    '-- roles come from externaldata/roledata.xlsx.',
    '',
    'truncate public.team_members, public.team_nodes restart identity cascade;',
    '',
    '-- nodes (parents inserted before children: divisions, then departments,',
    '-- then roles — the generation order already satisfies this).',
    'insert into public.team_nodes (id, parent_id, name, kind, position) values',
]
node_vals = []
for n in nodes:
    node_vals.append(
        f"  ({sql_str(n['id'])}::uuid, "
        f"{('null' if n['parent_id'] is None else sql_str(n['parent_id']) + '::uuid')}, "
        f"{sql_str(n['name'])}, {sql_str(n['kind'])}, {n['position']})")
lines.append(',\n'.join(node_vals) + ';')
lines.append('')
lines.append('insert into public.team_members (id, node_id, position, kkumail, '
             'prefix, full_name, nickname, student_id, year, major, confirmed) values')
mem_vals = []
for m in members:
    mem_vals.append(
        f"  ({sql_str(m['id'])}::uuid, {sql_str(m['node_id'])}::uuid, {m['position']}, "
        f"{sql_str(m['kkumail'])}, {sql_str(m['prefix'])}, {sql_str(m['full_name'])}, "
        f"{sql_str(m['nickname'])}, {sql_str(m['student_id'])}, {sql_str(m['year'])}, "
        f"{sql_str(m['major'])}, {sql_str(m['confirmed'])})")
lines.append(',\n'.join(mem_vals) + ';')
lines.append('')

with open(OUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print('nodes:', len(nodes), '| members:', len(members),
      '| loose role nodes:', stats['loose_nodes'])
print('people:', stats['people'], '| member rows:', stats['member_rows'])
print('wrote', OUT_JSON)
print('wrote', OUT_SQL)
